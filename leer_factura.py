import pdfplumber
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import tkinter as tk
from tkinter import filedialog

def extract_invoice_data(pdf_path):
    all_data = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            lines = text.split('\n')
            in_products_section = False
            current_product = None
            
            for line in lines:
                if 'Código Referencia Descripción' in line:
                    in_products_section = True
                    continue
                if any(end_marker in line for end_marker in ['Empresa adherida', 'Forma de Pago', 'CONDICIONES DE VENTA']):
                    in_products_section = False
                    if current_product:
                        all_data.append(current_product)
                        current_product = None
                    continue
                if in_products_section and line.strip():
                    pattern = r'^(\d+)\s+([^\s]+)\s+(.+?)\s+(\d+)\s+([\d,]+)\s+([\d,]+)\s+(\d+)$'
                    match = re.match(pattern, line.strip())
                    if match:
                        if current_product:
                            all_data.append(current_product)
                        codigo, referencia, descripcion, cantidad, precio, importe, iva = match.groups()
                        current_product = {
                            'Código': codigo,
                            'Referencia': referencia.strip(),
                            'Descripción': descripcion.strip(),
                            'Cantidad': cantidad,
                            'Precio': precio,
                            'IVA': iva
                        }
                    elif current_product:
                        if not re.search(r'\d+[\s,]+\d+[\s,]+\d+$', line.strip()):
                            current_product['Descripción'] += " " + line.strip()
                        else:
                            all_data.append(current_product)
                            current_product = None
            if current_product:
                all_data.append(current_product)
    
    cleaned_data = []
    for item in all_data:
        try:
            descripcion = re.sub(r'\s+', ' ', item['Descripción']).strip()
            cleaned_data.append({
                'Código': item['Código'],
                'Referencia': item['Referencia'],
                'Descripción': descripcion,
                'Cantidad': int(item['Cantidad']),
                'Precio': float(item['Precio'].replace(',', '.')),
                'IVA': int(item['IVA'])
            })
        except Exception as e:
            print(f"Error procesando item: {item}, error: {e}")
            continue
    return pd.DataFrame(cleaned_data)

def clean_descriptions(df):
    description_map = {
        'AGENDA ESCOLAR LIDERPAPEL 25-26 ESPIRAL': {
            'LP541': 'AGENDA ESCOLAR LIDERPAPEL 25-26 ESPIRAL BASIC AZUL DIA PAGINA',
            'LP542': 'AGENDA ESCOLAR LIDERPAPEL 25-26 ESPIRAL BASIC NEGRO DIA PAGINA',
            'LP543': 'AGENDA ESCOLAR LIDERPAPEL 25-26 ESPIRAL BASIC ROSA DIA PAGINA',
            'LP544': 'AGENDA ESCOLAR LIDERPAPEL 25-26 ESPIRAL BASIC ROJO DIA PAGINA',
            'LP546': 'AGENDA ESCOLAR LIDERPAPEL 25-26 ESPIRAL BASIC VERDE MENTA DIA PAGINA'
        },
        'BOLIGRAFO': {
            '8373602': 'BOLIGRAFO BIC CRISTAL ORIGINAL TINTA AZUL',
            'KF18625': 'BOLIGRAFO Q-CONNECT RETRACTIL BORRABLE 0,7 MM COLOR AZUL',
            'KF18626': 'BOLIGRAFO Q-CONNECT RETRACTIL BORRABLE 0,7 MM COLOR ROJO'
        }
    }
    for idx, row in df.iterrows():
        desc, ref = row['Descripción'], row['Referencia']
        if 'AGENDA ESCOLAR LIDERPAPEL 25-26 ESPIRAL' in desc and ref in description_map['AGENDA ESCOLAR LIDERPAPEL 25-26 ESPIRAL']:
            df.at[idx, 'Descripción'] = description_map['AGENDA ESCOLAR LIDERPAPEL 25-26 ESPIRAL'][ref]
        elif 'BOLIGRAFO' in desc and ref in description_map['BOLIGRAFO']:
            df.at[idx, 'Descripción'] = description_map['BOLIGRAFO'][ref]
    return df

def save_to_excel(df, pdf_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Factura"
    
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    column_widths = {'A': 10, 'B': 15, 'C': 60, 'D': 10, 'E': 10, 'F': 8}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                if cell.column in [4, 5]:
                    cell.alignment = Alignment(horizontal='right')
                elif cell.column == 6:
                    cell.alignment = Alignment(horizontal='center')
    
    total_row = len(df) + 3
    ws.cell(row=total_row, column=3, value="TOTAL PRODUCTOS:").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=len(df)).font = Font(bold=True)
    ws.cell(row=total_row+1, column=3, value="TOTAL UNIDADES:").font = Font(bold=True)
    ws.cell(row=total_row+1, column=4, value=df['Cantidad'].sum()).font = Font(bold=True)
    ws.cell(row=total_row+2, column=3, value="VALOR TOTAL:").font = Font(bold=True)
    total_value = (df['Cantidad'] * df['Precio']).sum()
    ws.cell(row=total_row+2, column=4, value=total_value).font = Font(bold=True)
    ws.cell(row=total_row+2, column=4).number_format = '#,##0.00€'
    
    excel_path = os.path.splitext(pdf_path)[0] + ".xlsx"
    wb.save(excel_path)
    return excel_path

def process_invoice_pdf(pdf_path):
    print(f"Procesando: {pdf_path}")
    df = extract_invoice_data(pdf_path)
    if df.empty:
        print("⚠ No se pudieron extraer datos.")
        return None
    df = clean_descriptions(df)
    excel_path = save_to_excel(df, pdf_path)
    print(f"✓ Guardado: {excel_path}")
    return df

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    carpeta = filedialog.askdirectory(title="Selecciona la carpeta con PDFs")
    
    if carpeta:
        pdf_files = [f for f in os.listdir(carpeta) if f.lower().endswith(".pdf")]
        if not pdf_files:
            print("No se encontraron archivos PDF en la carpeta.")
        else:
            for pdf in pdf_files:
                pdf_path = os.path.join(carpeta, pdf)
                process_invoice_pdf(pdf_path)
    else:
        print("No se seleccionó ninguna carpeta.")
